import docx
import openpyxl
import os

def extract_docx(file_path):
    doc = docx.Document(file_path)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    return '\n'.join(full_text)

def extract_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    result = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        result.append(f"--- Sheet: {sheet_name} ---")
        for row in sheet.iter_rows(values_only=True):
            result.append('\t'.join([str(cell) if cell is not None else "" for cell in row]))
    return '\n'.join(result)

docx_path = r'c:\Users\Sheikh PC\StudioProjects\untitled3\deveopment_R\nijib BRD.docx'
xlsx_path = r'c:\Users\Sheikh PC\StudioProjects\untitled3\deveopment_R\Aditi-8 week workout plan MASTER.xlsx'

if os.path.exists(docx_path):
    brd_text = extract_docx(docx_path)
    with open(r'c:\Users\Sheikh PC\StudioProjects\untitled3\deveopment_R\BRD_extracted.txt', 'w', encoding='utf-8') as f:
        f.write(brd_text)
    print("Extracted BRD text.")

if os.path.exists(xlsx_path):
    workout_text = extract_xlsx(xlsx_path)
    with open(r'c:\Users\Sheikh PC\StudioProjects\untitled3\deveopment_R\workout_plan_extracted.txt', 'w', encoding='utf-8') as f:
        f.write(workout_text)
    print("Extracted workout plan text.")
